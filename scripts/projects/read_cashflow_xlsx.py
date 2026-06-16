"""
read_cashflow_xlsx.py
======================
Reads all 8 manual tabs from a cashflow xlsx and writes a separate normalized
JSON file per tab into data/projects/.

IMPORTANT — source of truth:
  The xlsx this parser reads is downloaded fresh from the live Google Sheet
  `FirstRain-Cashflow-Master` by /finance Step 3a (via Drive MCP
  `download_file_content` → base64 decode → cache file). Sonal edits the
  Google Sheet, never the xlsx on disk. This script never reads from
  ~/Desktop — that location only holds the static template/schema reference.

The 2 auto-filled tabs (Bank_Transactions, Statutory_Filings) are handled
by separate Gmail parsers — this file only handles what Sonal/Niloy type by
hand into the sheet.

Tabs handled here:
  1. Cash_Position       → data/projects/sheet_cash_position.json
  2. Treasury_Holdings   → data/projects/sheet_treasury.json
  3. Receivables         → data/projects/sheet_receivables.json
  4. Payables            → data/projects/sheet_payables.json
  5. Statutory           → data/projects/sheet_statutory.json (manual schedule)
  6. Notes               → data/projects/sheet_notes.json
  7. Projects            → data/projects/sheet_projects.json

Source priority:
  1. CLI arg (path to xlsx)
  2. $FR_CASHFLOW_XLSX env var
  3. data/projects/_cache/cashflow_master.xlsx (default — populated by /finance)

Usage:
    python scripts/projects/read_cashflow_xlsx.py
    python scripts/projects/read_cashflow_xlsx.py path/to/custom.xlsx

Exits 0 if all 8 tabs parsed without fatal errors; any per-row warnings are
captured in the respective JSON's meta.warnings list.
Exits 3 with a "Sheet not yet migrated" message if any required tab is missing
(typical when Sonal hasn't replaced the old Sheet structure yet).
"""
from __future__ import annotations

import json
import os
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# ---------- Paths ----------
# DEFAULT is the cache location populated by /finance Step 3a (Drive MCP download).
# The Desktop xlsx is intentionally NOT read here — that's the schema template only.
DEFAULT_XLSX = Path("data/projects/_cache/cashflow_master.xlsx")
OUTPUT_DIR = Path("data/projects")
REQUIRED_TABS = {
    "Cash_Position", "Treasury_Holdings", "Receivables", "Payables",
    "Notes", "Projects",
}
# Statutory tab is accepted as either name (parser handles both via _load_tab fallback)
STATUTORY_TAB_NAMES = {"Statutory", "Statutory-1"}

# ---------- Shared helpers ----------

def _to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("₹", "").replace(",", "").strip()
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def _to_int(v: Any) -> int:
    return int(_to_float(v))


def _parse_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _clean(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    s = _clean(v).lower()
    return s in ("true", "yes", "1", "y", "t")


def _normalize_header(h: Any) -> str:
    """Normalize a header cell to lower_snake_case.
    Handles camelCase ('obligationType' → 'obligation_type'), spaces,
    and stray whitespace. Idempotent on already-snake headers.
    """
    s = _clean(h)
    if not s:
        return ""
    # Insert underscore before each uppercase letter that follows a lowercase
    # letter or digit (camelCase → camel_case)
    out = []
    for i, ch in enumerate(s):
        if i > 0 and ch.isupper() and (s[i-1].islower() or s[i-1].isdigit()):
            out.append("_")
        out.append(ch)
    return "".join(out).lower().replace(" ", "_").replace("-", "_")


def _load_tab(wb, tab_name) -> tuple[list[str], list[list]]:
    """Return (headers, data_rows). Headers normalized to lower_snake_case.
    `tab_name` can be a string or a list/tuple of acceptable names — the
    first one that exists in the workbook wins.
    """
    candidates = [tab_name] if isinstance(tab_name, str) else list(tab_name)
    resolved = None
    for cand in candidates:
        if cand in wb.sheetnames:
            resolved = cand
            break
    if resolved is None:
        raise ValueError(
            f"None of {candidates} found in workbook. Available: {wb.sheetnames}"
        )
    ws = wb[resolved]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []
    headers = [_normalize_header(h) for h in all_rows[0]]
    data_rows = [list(r) for r in all_rows[1:] if any(c is not None and str(c).strip() for c in r)]
    return headers, data_rows


def _row_dict(headers: list[str], row: list) -> dict:
    padded = list(row) + [None] * (len(headers) - len(row))
    return dict(zip(headers, padded))


def _write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        json.dump(payload, f, indent=2, default=str)


def _meta(count: int, warnings: list) -> dict:
    return {
        "count": count,
        "read_at": datetime.now().astimezone().isoformat(),
        "warnings": warnings,
    }


# ---------- 1. Cash_Position ----------

def parse_cash_position(wb) -> dict:
    headers, rows = _load_tab(wb, "Cash_Position")
    if not rows:
        return {"cash": None, "history": [], "meta": _meta(0, ["tab_empty"])}

    history = []
    warnings = []
    for row in rows:
        d = _row_dict(headers, row)
        snap_date = _parse_date(d.get("date"))
        if not snap_date:
            warnings.append(f"unparseable_date:{d.get('date')}")
            continue
        history.append({
            "date": snap_date.isoformat(),
            "operatingCash": _to_float(d.get("hdfc_operating")),
            "hdfcLast4": _clean(d.get("hdfc_account_last4")),
            "treasury": _to_float(d.get("treasury_total")),
            "treasuryLastRefreshed": (_parse_date(d.get("treasury_last_refreshed")) or snap_date).isoformat(),
            "odUtilized": _to_float(d.get("od_utilized")),
            "odLimit": _to_float(d.get("od_limit")),
            "totalReceivables": _to_float(d.get("total_receivables")),
            "totalPayables": _to_float(d.get("total_payables")),
            "updatedBy": _clean(d.get("updated_by")),
            "notes": _clean(d.get("notes")),
        })

    history.sort(key=lambda h: h["date"], reverse=True)
    latest = history[0] if history else None

    # Alert flags — Niloy's floors from CLAUDE.md
    alerts = []
    if latest:
        if latest["operatingCash"] < 7650000:
            alerts.append(f"OPERATING_CASH_BELOW_FLOOR: ₹{latest['operatingCash']:,.0f} < ₹76,50,000")
        if latest["odUtilized"] > 0:
            alerts.append(f"OD_UTILIZED: ₹{latest['odUtilized']:,.0f} of ₹{latest['odLimit']:,.0f}")

    return {
        "cash": latest,
        "history": history,
        "alerts": alerts,
        "meta": _meta(len(history), warnings),
    }


# ---------- 2. Treasury_Holdings ----------

VALID_INVESTMENT_TYPES = {"liquid", "debt_short", "debt_long", "equity", "hybrid", "fd"}

def parse_treasury(wb) -> dict:
    headers, rows = _load_tab(wb, "Treasury_Holdings")
    holdings = []
    warnings = []
    by_type: dict[str, float] = defaultdict(float)

    for row in rows:
        d = _row_dict(headers, row)
        name = _clean(d.get("instrument_name"))
        if not name or name.upper() == "TOTAL":
            continue
        inv_type = _clean(d.get("investment_type")).lower()
        if inv_type and inv_type not in VALID_INVESTMENT_TYPES:
            warnings.append(f"invalid_investment_type:{inv_type}:{name}")
        notes_val = _clean(d.get("notes")) or ""
        if "redeemed" in notes_val.lower():
            continue
        current_val = _to_float(d.get("current_value"))
        h = {
            "instrument": name,
            "investmentType": inv_type,
            "folio": _clean(d.get("folio_no")),
            "amc": _clean(d.get("amc")),
            "units": _to_float(d.get("units")),
            "nav": _to_float(d.get("nav")),
            "currentValue": current_val,
            "investedValue": _to_float(d.get("invested_value")),
            "unrealizedGain": _to_float(d.get("unrealized_gain")),
            "exitLoadPct": _to_float(d.get("exit_load_pct")),
            "exitLoadValidUntil": (_parse_date(d.get("exit_load_valid_until")) or None),
            "lockInUntil": (_parse_date(d.get("lock_in_until")) or None),
            "ltcgApplicable": _to_bool(d.get("ltcg_applicable")),
            "lastStatementDate": (_parse_date(d.get("last_statement_date")) or None),
            "notes": notes_val,
        }
        # Normalize dates to ISO strings
        for k in ("exitLoadValidUntil", "lockInUntil", "lastStatementDate"):
            h[k] = h[k].isoformat() if h[k] else None
        holdings.append(h)
        by_type[inv_type or "unknown"] += current_val

    total = sum(by_type.values())
    return {
        "holdings": holdings,
        "totals": {
            "total": total,
            "byType": dict(by_type),
        },
        "meta": _meta(len(holdings), warnings),
    }


# ---------- 3. Receivables ----------

def parse_receivables(wb) -> dict:
    headers, rows = _load_tab(wb, "Receivables")
    today = date.today()
    receivables = []
    warnings = []
    total_balance = 0.0
    overdue_count = 0

    for row in rows:
        d = _row_dict(headers, row)
        client = _clean(d.get("client_name"))
        if not client or client.upper() == "TOTAL":
            continue
        due = _parse_date(d.get("due_date"))
        expected = _parse_date(d.get("expected_date"))
        # Use expected_date when it's later than due_date (e.g. show closes
        # after invoice date, so payment can't realistically arrive on
        # invoice due_date). This matches how the exec team tracks
        # collections and stops false 50d-overdue alerts for invoices
        # that are actually on schedule.
        ref_date = expected if (expected and (not due or expected > due)) else due
        days_od = max(0, (today - ref_date).days) if ref_date else 0
        balance = _to_float(d.get("balance"))
        if balance > 0:
            total_balance += balance
            if days_od > 0:
                overdue_count += 1

        receivables.append({
            "client": client,
            "invoiceNo": _clean(d.get("tally_invoice_no")),
            "invoiceDate": (_parse_date(d.get("invoice_date")) or None),
            "showProject": _clean(d.get("show_project")),
            "spTotal": _to_float(d.get("sp_total")),
            "gstAmount": _to_float(d.get("gst_amount")),
            "invoiceTotal": _to_float(d.get("invoice_total")),
            "receivedToDate": _to_float(d.get("received_todate")),
            "balance": balance,
            "dueDate": due.isoformat() if due else None,
            "daysOverdue": days_od,
            "status": _clean(d.get("status")),
            "execOwner": _clean(d.get("exec_owner")),
            "priority": _clean(d.get("priority")),
            "lastContactDate": (_parse_date(d.get("last_contact_date")) or None),
            "lastContactType": _clean(d.get("last_contact_type")),
            "expectedDate": expected.isoformat() if expected else None,
            "clientNotes": _clean(d.get("client_notes")),
        })

    # Stringify all date fields
    for r in receivables:
        if isinstance(r.get("invoiceDate"), date):
            r["invoiceDate"] = r["invoiceDate"].isoformat()
        if isinstance(r.get("lastContactDate"), date):
            r["lastContactDate"] = r["lastContactDate"].isoformat()

    return {
        "receivables": receivables,
        "totals": {
            "totalBalance": total_balance,
            "overdueCount": overdue_count,
            "count": len(receivables),
        },
        "meta": _meta(len(receivables), warnings),
    }


# ---------- 4. Payables ----------

VALID_FLEXIBILITY = {"must-pay", "can-delay-7d", "can-delay-30d", "can-delay-45d", "can-delay-60d", "can-delay-90d"}

def parse_payables(wb) -> dict:
    headers, rows = _load_tab(wb, "Payables")
    today = date.today()
    payables = []
    warnings = []
    total_balance = 0.0
    due_this_week = 0.0

    for row in rows:
        d = _row_dict(headers, row)
        vendor = _clean(d.get("vendor_name"))
        if not vendor or vendor.upper() == "TOTAL":
            continue
        due = _parse_date(d.get("due_date"))
        days_to_due = (due - today).days if due else None
        balance = _to_float(d.get("balance"))
        status = _clean(d.get("status")).lower()

        flexibility = _clean(d.get("flexibility")).lower()
        if flexibility and flexibility not in VALID_FLEXIBILITY:
            warnings.append(f"invalid_flexibility:{flexibility}:{vendor}")

        if balance > 0 and status != "paid":
            total_balance += balance
            if days_to_due is not None and 0 <= days_to_due <= 7:
                due_this_week += balance

        payables.append({
            "vendor": vendor,
            "invoiceNo": _clean(d.get("invoice_no")),
            "invoiceDate": (_parse_date(d.get("invoice_date")) or None),
            "invoiceAmount": _to_float(d.get("invoice_amount")),
            "gstAmount": _to_float(d.get("gst_amount")),
            "totalAmount": _to_float(d.get("total_amount")),
            "amountPaid": _to_float(d.get("amount_paid")),
            "balance": balance,
            "dueDate": due.isoformat() if due else None,
            "daysToDue": days_to_due,
            "status": status or "pending",
            "paymentTerms": _clean(d.get("payment_terms")),
            "priority": _clean(d.get("priority")),
            "flexibility": flexibility,
            "flexNote": _clean(d.get("flex_note")),
            "approver": _clean(d.get("approver")),
            "notes": _clean(d.get("notes")),
        })

    for p in payables:
        if isinstance(p.get("invoiceDate"), date):
            p["invoiceDate"] = p["invoiceDate"].isoformat()

    return {
        "payables": payables,
        "totals": {
            "totalBalance": total_balance,
            "dueThisWeek": due_this_week,
            "count": len(payables),
        },
        "meta": _meta(len(payables), warnings),
    }


# ---------- 5. Statutory (manual schedule — Gmail parser fills `filings` separately) ----------

def parse_statutory_manual(wb) -> dict:
    # Try the new tab first ("Statutory-1" with expanded schema) and fall back
    # to the original "Statutory" tab.
    headers, rows = _load_tab(wb, ["Statutory-1", "Statutory"])
    today = date.today()
    obligations = []
    overdue = []
    upcoming_30d = []
    warnings = []

    for row in rows:
        d = _row_dict(headers, row)
        ob_type = _clean(d.get("obligation_type"))
        if not ob_type or ob_type.upper() == "TOTAL":
            continue
        due = _parse_date(d.get("due_date"))
        filed = _parse_date(d.get("filed_date"))
        status = _clean(d.get("status")).lower()
        days_to_due = (due - today).days if due else None

        # Effective net payable.
        # The sheet's `amountPayable` = taxPayableSales − creditInputPurchase, but
        # is only filled once a return is finalized/filed. For not-yet-filed
        # ("projected") rows it stays 0 even when a real gross tax-on-sales
        # estimate exists (e.g. May GSTR-3B: taxPayableSales ₹17.9L, input credit
        # not yet entered → net ₹17.9L). Showing 0 would hide a material cash
        # obligation in both the dashboard and the monthly cash-flow projection.
        # So: prefer the finalized amountPayable; otherwise fall back to the net
        # estimate (tax on sales − input credit) when that is positive.
        amt_final = _to_float(d.get("amount_payable"))
        tax_sales = _to_float(d.get("tax_payable_sales"))
        credit_in = _to_float(d.get("credit_input_purchase"))
        net_est = round(tax_sales - credit_in, 2)
        effective_payable = amt_final if amt_final > 0 else (net_est if net_est > 0 else 0.0)
        is_estimate = amt_final <= 0 and net_est > 0

        entry = {
            "obligationType": ob_type,
            "periodFor": _clean(d.get("period_for")),
            "dueDate": due.isoformat() if due else None,
            "amountPayable": effective_payable,
            "amountPayableFinal": amt_final,   # raw sheet value (0 until filed)
            "taxPayableSales": tax_sales,
            "creditInputPurchase": credit_in,
            "netEstimate": net_est,
            "cgstNet": _to_float(d.get("cgst_net")),
            "sgstNet": _to_float(d.get("sgst_net")),
            "igstNet": _to_float(d.get("igst_net")),
            "isEstimate": is_estimate,
            "filedDate": filed.isoformat() if filed else None,
            "challanRef": _clean(d.get("challan_ref")),
            "status": status or "pending",
            "lateFee": _to_float(d.get("late_fee_if_any")),
            "filingPortal": _clean(d.get("filing_portal")),
            "responsible": _clean(d.get("responsible")),
            "notes": _clean(d.get("notes")),
            "daysToDue": days_to_due,
        }
        obligations.append(entry)

        if status != "filed" and due and due < today:
            overdue.append(entry)
        elif status != "filed" and days_to_due is not None and 0 <= days_to_due <= 30:
            upcoming_30d.append(entry)

    return {
        "obligations": obligations,
        "overdue": overdue,
        "upcoming30d": upcoming_30d,
        "meta": _meta(len(obligations), warnings),
    }


# ---------- 7. Notes (dated entries + monthly totals for cash-flow extrapolation) ----------

def parse_notes(wb) -> dict:
    headers, rows = _load_tab(wb, "Notes")
    entries = []
    monthly_revenue: dict[str, float] = {}
    monthly_burn: dict[str, float] = {}
    warnings = []

    for row in rows:
        d = _row_dict(headers, row)
        category = _clean(d.get("category"))
        entry_text = _clean(d.get("entry"))
        if not category and not entry_text:
            continue

        # Monthly revenue overrides: Monthly_Revenue_YYYY-MM
        if category.upper().startswith("MONTHLY_REVENUE_"):
            month = category[len("Monthly_Revenue_"):]
            monthly_revenue[month] = _to_float(entry_text)
            continue

        # Per-month burn overrides: CONFIG_MONTHLY_BURN_YYYY-MM (e.g. CONFIG_MONTHLY_BURN_2026-05)
        if category.upper().startswith("CONFIG_MONTHLY_BURN_"):
            month = category[len("CONFIG_MONTHLY_BURN_"):]
            monthly_burn[month] = _to_float(entry_text)
            continue

        entry_date = _parse_date(d.get("date"))
        entries.append({
            "date": entry_date.isoformat() if entry_date else None,
            "category": category,
            "entry": entry_text,
            "by": _clean(d.get("by")),
        })

    return {
        "entries": entries,
        "monthlyRevenue": monthly_revenue,
        "monthlyBurn": monthly_burn,
        "meta": _meta(len(entries) + len(monthly_revenue) + len(monthly_burn), warnings),
    }


# ---------- 8. Projects ----------

def parse_projects(wb) -> dict:
    headers, rows = _load_tab(wb, "Projects")
    projects = []
    warnings = []
    totals = {"sp": 0.0, "cp": 0.0, "commission": 0.0, "contribution": 0.0, "sizeSqm": 0.0}

    for row in rows:
        d = _row_dict(headers, row)
        company = _clean(d.get("company"))
        if not company or company.upper() == "TOTAL":
            continue
        sp = _to_float(d.get("sales_sp"))
        cp = _to_float(d.get("variable_cost_cp"))
        comm = _to_float(d.get("commission"))
        contribution = sp - cp - comm
        contrib_pct = (contribution / sp) if sp else 0.0
        totals["sp"] += sp
        totals["cp"] += cp
        totals["commission"] += comm
        totals["contribution"] += contribution
        totals["sizeSqm"] += _to_float(d.get("size_sqm"))

        # Margin floor check (CLAUDE.md): 33% India / 38% International
        region = _clean(d.get("region")) or "India"
        floor = 0.38 if region.lower() != "india" else 0.33
        below_floor = contrib_pct < floor

        projects.append({
            "company": company,
            "project": _clean(d.get("project")),
            "delivery_month": _clean(d.get("delivery_month")),
            "sizeSqm": _to_float(d.get("size_sqm")),
            "city": _clean(d.get("city")),
            "fabricator": _clean(d.get("fabricator")),
            "sales_sp": sp,
            "variable_cost_cp": cp,
            "commission": comm,
            "status": _clean(d.get("status")),
            "is_repeat": _to_bool(d.get("is_repeat")),
            "region": region,
            "exec": _clean(d.get("exec")) or "ND",
            "contribution": contribution,
            "contributionPct": contrib_pct,
            "belowMarginFloor": below_floor,
        })

    totals["contributionPct"] = (totals["contribution"] / totals["sp"]) if totals["sp"] else 0.0
    return {
        "projects": projects,
        "totals": totals,
        "meta": _meta(len(projects), warnings),
    }


# ---------- Orchestrator ----------

TAB_PARSERS = [
    ("Cash_Position",      "sheet_cash_position.json",   parse_cash_position),
    ("Treasury_Holdings",  "sheet_treasury.json",         parse_treasury),
    ("Receivables",        "sheet_receivables.json",      parse_receivables),
    ("Payables",           "sheet_payables.json",         parse_payables),
    ("Statutory",          "sheet_statutory.json",        parse_statutory_manual),
    ("Notes",              "sheet_notes.json",            parse_notes),
    ("Projects",           "sheet_projects.json",         parse_projects),
]


def main(xlsx_path: Path | None = None) -> int:
    path = xlsx_path or Path(os.environ.get("FR_CASHFLOW_XLSX", str(DEFAULT_XLSX)))
    if not path.exists():
        print(f"ERROR: xlsx not found at {path}", file=sys.stderr)
        print(
            "Hint: /finance Step 3a downloads the live Sheet to "
            "data/projects/_cache/cashflow_master.xlsx via Drive MCP. "
            "Run that step first.",
            file=sys.stderr,
        )
        return 2

    print(f"Reading {path}")
    wb = load_workbook(path, data_only=True)

    # Pre-flight: confirm the live Sheet has the new structure (post-migration).
    actual_tabs = set(wb.sheetnames)
    missing = REQUIRED_TABS - actual_tabs
    if not (actual_tabs & STATUTORY_TAB_NAMES):
        missing.add("Statutory (or Statutory-1)")
    if missing:
        print(
            "ERROR: live Sheet is missing required tabs — "
            "Sonal/Niloy hasn't completed the FirstRain-Cashflow-Master "
            "import yet.",
            file=sys.stderr,
        )
        print(f"  Found tabs:    {sorted(actual_tabs)}", file=sys.stderr)
        print(f"  Missing tabs:  {sorted(missing)}", file=sys.stderr)
        print(
            "Fix: In Google Sheets, File → Import → Upload "
            "~/Desktop/FirstRain-Cashflow-Master.xlsx → Replace spreadsheet.",
            file=sys.stderr,
        )
        return 3

    summary = []
    for tab_name, out_filename, parser in TAB_PARSERS:
        try:
            result = parser(wb)
            out_path = OUTPUT_DIR / out_filename
            _write_json(out_path, result)
            count = result.get("meta", {}).get("count", 0)
            warn_count = len(result.get("meta", {}).get("warnings", []))
            summary.append((tab_name, count, warn_count, out_path.name))
            print(f"  ✓ {tab_name:20s} → {out_path.name:35s} ({count} rows, {warn_count} warnings)")
        except Exception as e:
            print(f"  ✗ {tab_name}: {e}", file=sys.stderr)
            summary.append((tab_name, 0, -1, f"FAILED: {e}"))

    print("\nSummary:")
    for tab, count, warnings, out in summary:
        flag = "✓" if warnings >= 0 else "✗"
        print(f"  {flag} {tab:20s} {count:4d} rows · {max(0,warnings)} warnings · {out}")

    failed = [s for s in summary if s[2] < 0]
    return 1 if failed else 0


if __name__ == "__main__":
    path_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    sys.exit(main(path_arg))
